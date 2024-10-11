#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Copyright (c) 2023 Huawei Device Co., Ltd.
Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.

Description: Use ark to execute workload test suite
"""

import argparse
import datetime
import logging
import os
import sys
import xml.etree.ElementTree as ET
from collections import namedtuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment


class Constants:
    DETERIORATION_BOUNDARY_VALUE = 0.05
    RET_OK = 0
    RET_ERROR = 1
    logger = None


def get_logger(logger_name, log_file_path, level=logging.INFO):
    formatter = logging.Formatter(fmt='[%(asctime)s]  [%(levelname)s]   %(message)s',
                                  datefmt='%Y-%b-%d  %H:%M:%S')

    fh = logging.FileHandler(encoding='utf-8', mode='a', filename=log_file_path)
    fh.setFormatter(formatter)
    fh.setLevel(logging.DEBUG)
    ch = logging.StreamHandler()
    ch.setFormatter(formatter)
    ch.setLevel(logging.DEBUG)
    log = logging.getLogger(logger_name)
    log.addHandler(fh)
    log.addHandler(ch)
    log.setLevel(level)

    return log


def get_today_test_result(xml_folder_path):
    xml_list = []
    test_retult_list = []
    for root, _, files in os.walk(xml_folder_path):
        for file in files:
            file_path = os.path.join(root, file)
            xml_list.append(file_path)

    sorted_xml_list = sorted(xml_list)
    CaseTestDataType = namedtuple('test', ['result', 'exec_time'])
    for _, xml_file in enumerate(sorted_xml_list):
        test_result = {}
        tree = ET.parse(xml_file)
        root = tree.getroot()
        for testsuite in root.findall('testsuite'):
            if testsuite.get('name') != "JSNApiSplTest":
                continue
            # ergodic testcase
            for testcase in testsuite.findall('testcase'):
                time_str = testcase.get('time')
                time = float(time_str.strip())
                case_test_data = CaseTestDataType(testcase.get('result'), time)
                test_result[testcase.get('name')] = case_test_data
            test_retult_list.append(test_result)

    return test_retult_list


def get_yestoday_test_data(yestoday_daily_report):
    test_data = {}
    try:
        wb = load_workbook(yestoday_daily_report)
        ws = wb.worksheets[0]
    except FileNotFoundError:
        return test_data
    CaseTestDataType = namedtuple('test', ['result', 'exec_time'])
    for row_num in range(2, ws.max_row + 1):
        js_case_name_tmp = str(ws.cell(row=row_num, column=1).value)
        excu_status = str(ws.cell(row=row_num, column=2).value)
        excu_time = str(ws.cell(row=row_num, column=8).value)
        case_data = CaseTestDataType(excu_status, excu_time)
        test_data[js_case_name_tmp] = case_data

    return test_data


def get_yesday_average_time(yestoday_data, workload_case_name):
    yesday_average_cost_time = "0"
    if len(yestoday_data) > 0:
        for key in yestoday_data:
            if workload_case_name in key:
                yesday_average_cost_time = str(yestoday_data[workload_case_name].exec_time)
                break

    return float(yesday_average_cost_time)


def generate_daily_report(daily_report_file, today_data, yestoday_data):
    '''
        report_file, daily
        today_data, list. element is dictionary. include five cost times info.value of dict includes the info of
            excute status and exute_time
        yestoday_data: dictionary. values of dict are yesterday workload use cases excuting average cost times
    '''
    Constants.logger.info("begin to generate report.")
    if len(today_data) == 0:
        Constants.logger.error("no today testing data, please check it!.")
        return Constants.RET_ERROR
    wb = load_workbook(daily_report_file)
    ws = wb.worksheets[0]
    for workload_case_name in today_data[0].keys():
        notes = ' '
        first_cost_time = today_data[0][workload_case_name].exec_time
        if today_data[0][workload_case_name].result == 'completed':
            excute_status = 'pass'
        else:
            excute_status = today_data[0][workload_case_name].result
        second_cost_time = today_data[1][workload_case_name].exec_time
        third_cost_time = today_data[2][workload_case_name].exec_time
        fourth_cost_time = today_data[3][workload_case_name].exec_time
        fifth_cost_time = today_data[4][workload_case_name].exec_time
        time_list = [first_cost_time, second_cost_time, third_cost_time, third_cost_time, fourth_cost_time,
                     fourth_cost_time, fifth_cost_time]
        time_list_len = len(time_list)
        if time_list_len == 0:
            today_average_cost_time = 0
        else:
            today_average_cost_time = sum(time_list) / time_list_len
        yesday_average_cost_time = get_yesday_average_time(yestoday_data, workload_case_name)

        try:
            tmp = today_average_cost_time / yesday_average_cost_time
            if tmp >= (1.0 + Constants.DETERIORATION_BOUNDARY_VALUE):
                is_degraded_str = str(True)
            else:
                is_degraded_str = str(False)
        except ZeroDivisionError:
            is_degraded_str = 'NA'

        new_row = [workload_case_name, excute_status, first_cost_time, second_cost_time, third_cost_time,
                   fourth_cost_time, fifth_cost_time, today_average_cost_time, yesday_average_cost_time,
                   is_degraded_str, notes]

        ws.append(new_row)
        if is_degraded_str == str(True):
            ws.cell(row=ws.max_row, column=10).fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                                                  fill_type='solid')
    wb.save(daily_report_file)
    Constants.logger.info("generate report successfully with no summary infomation.")
    return Constants.RET_OK


def get_degraded_num(daily_report_file):
    wb = load_workbook(daily_report_file)
    ws = wb.worksheets[0]
    degraded_num = 0
    for row_num in range(2, ws.max_row + 1):
        is_degraded = str(ws.cell(row=row_num, column=10).value)
        if is_degraded == str(True):
            degraded_num += 1
    return degraded_num


def get_summary_info(xml_folder_path, daily_report_file):
    local_summary_data_dict = {}
    xml_list = []
    for root, _, files in os.walk(xml_folder_path):
        for file in files:
            file_ext = os.path.splitext(file)[1]
            if file_ext == '.xml':
                file_path = os.path.join(root, file)
                xml_list.append(file_path)

    sorted_xml_list = sorted(xml_list)

    # calculate average excute cost time
    excute_time_list = []
    for _, xml in enumerate(sorted_xml_list):
        tmp_tree = ET.parse(xml)
        tmp_root = tmp_tree.getroot()
        for testsuite in tmp_root.findall('testsuite'):
            if testsuite.get('name') != "JSNApiSplTest":
                continue
            time_str = testsuite.get('time')
            time = float(time_str.strip())
            excute_time_list.append(time)
            break
    num = len(excute_time_list)
    if num == 0:
        today_average_excute_time = 0
    else:
        today_average_excute_time = sum(excute_time_list) / num
    tree = ET.parse(sorted_xml_list[0])
    root = tree.getroot()
    for testsuite in root.findall('testsuite'):
        if testsuite.get('name') != "JSNApiSplTest":
            continue

        total_count = int(testsuite.get('tests')) + int(testsuite.get('failures')) + int(testsuite.get('disabled')) + \
            int(testsuite.get('skipped')) + int(testsuite.get('errors'))
        local_summary_data_dict['total_count'] = total_count
        local_summary_data_dict['pass_count'] = testsuite.get('tests')
        local_summary_data_dict['failures_count'] = testsuite.get('failures')
        local_summary_data_dict['disabled_count'] = testsuite.get('disabled')
        local_summary_data_dict['skipped_count'] = testsuite.get('skipped')
        local_summary_data_dict['errors_count'] = testsuite.get('errors')

        local_summary_data_dict['date'] = testsuite.get('timestamp')
        break

    local_summary_data_dict['degraded_num'] = get_degraded_num(daily_report_file)
    local_summary_data_dict["today_average_excute_time"] = today_average_excute_time
    local_summary_data_dict["degraded percentage upper limit"] = Constants.DETERIORATION_BOUNDARY_VALUE
    return local_summary_data_dict


def init_percentage_varibales_1(summary_data):
    fixed_str = "0%"
    percentage = fixed_str
    percentage1 = fixed_str
    percentage2 = fixed_str
    total_num = int(summary_data['total_count'])
    if total_num > 0:
        percentage = str(round((int(summary_data["pass_count"]) / total_num) * 100, 2)) + '%'
        percentage1 = str(round((int(summary_data["failures_count"]) / total_num) * 100, 2)) + '%'
        percentage2 = str(round((int(summary_data["disabled_count"]) / total_num) * 100, 2)) + '%'

    return percentage, percentage1, percentage2


def init_percentage_varibales_2(summary_data):
    fixed_str = "0%"
    percentage3 = fixed_str
    percentage4 = fixed_str
    percentage5 = fixed_str

    total_num = int(summary_data['total_count'])
    if total_num > 0:
        percentage3 = str(round((int(summary_data["skipped_count"]) / total_num) * 100, 2)) + '%'
        percentage4 = str(round((int(summary_data["errors_count"]) / total_num) * 100, 2)) + '%'
        percentage5 = str(round((int(summary_data["degraded_num"]) / total_num) * 100, 2)) + '%'
    return percentage3, percentage4, percentage5


def append_summary_info(daily_report_file, summary_data):
    """
        summary info:
            totle count:        percentage:
            pass count:         percentage:
            failures count:     percentage:
            disabled count:     percentage:
            skipped count:      percentage:
            errors count:       percentage:
            degraded count:     percentage:
            today average excute time(s):
            degraded percentage upper limit:
            date:
    """
    Constants.logger.info("begin to append summary infomation to  today report.")
    wb = load_workbook(daily_report_file)
    ws = wb.worksheets[0]
    percentage_str = 'percentage:'
    # append 3 blank
    blank_num = 3
    for _ in range(blank_num):
        new_row = [' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
        ws.append(new_row)

    new_row = ['summary info:', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
    ws.append(new_row)

    total_num = int(summary_data['total_count'])
    percentage, percentage1, percentage2 = init_percentage_varibales_1(summary_data)
    percentage3, percentage4, percentage5 = init_percentage_varibales_2(summary_data)
    new_row = ['totle count:', total_num, ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
    ws.append(new_row)

    new_row = ['pass count:', summary_data["pass_count"], 'percentage:', percentage, ' ', ' ', ' ',
               ' ', ' ', ' ', ' ']
    ws.append(new_row)

    new_row = ['failures count:', summary_data["failures_count"], percentage_str, percentage1,
               ' ', ' ', ' ', ' ', ' ', ' ', ' ']
    ws.append(new_row)

    new_row = ['disabled count:', summary_data["disabled_count"], percentage_str, percentage2,
               ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
    ws.append(new_row)

    new_row = ['skipped count:', summary_data["skipped_count"], percentage_str, percentage3, ' ',
               ' ', ' ', ' ', ' ', ' ', ' ']
    ws.append(new_row)

    new_row = ['errors count:', summary_data["errors_count"], percentage_str, percentage4, ' ', ' ',
               ' ', ' ', ' ', ' ', ' ']
    ws.append(new_row)

    new_row = ['Degraded count:', int(summary_data["degraded_num"]), percentage_str,
               percentage5, ' ', ' ', ' ', ' ', ' ', ' ', ' ']
    ws.append(new_row)

    new_row = ['Today average excute time(s):', float(summary_data["today_average_excute_time"]), ' ', ' ', ' ', ' ',
               ' ', ' ', ' ', ' ', ' ']
    ws.append(new_row)

    new_row = ['Degraded percentage upper limit:', float(summary_data["degraded percentage upper limit"]), ' ', ' ',
               ' ', ' ', ' ', ' ', ' ', ' ', ' ']
    ws.append(new_row)

    new_row = ['Date:', summary_data["date"], ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
    ws.append(new_row)

    for i in range(2, ws.max_row + 1):
        align = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=0, text_rotation=0)
        ws.cell(row=i, column=2).alignment = align

    wb.save(daily_report_file)


def get_args():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--work_path",
        "-w",
        required=True,
        help="work folder",
    )

    parser.add_argument(
        "--rerport_folder_path",
        "-o",
        default=None,
        required=True,
        help="report save path, default current folder",
    )
    args = parser.parse_args()

    if not os.path.exists(args.work_path):
        Constants.logger.error("parameter --work is not exit. Please check it! work path: %s", args.work_path)
        raise RuntimeError(f"error bad  parameters  --work. work: {args.work_path}")

    if not os.path.isdir(args.work_path):
        Constants.logger.error("parameter --work is not folder. Please check it! work path: %s", args.work_path)
        raise RuntimeError(f"error bad  parameters  --work. work: {args.work_path}")

    return args


def init_report(daily_report_file):
    Constants.logger.info("begin to initialize today report")
    try:
        wb = load_workbook(daily_report_file)
        ws = wb.worksheets[0]

    except FileNotFoundError:
        headers_row = ['用例名称', '执行状态', '用例执行耗时-1(s)', '用例执行耗时-2(s)', '用例执行耗时-3(s)',
                       '用例执行耗时-4(s)', '用例执行耗时-5(s)', '当日用例平均耗时(s)', '昨日用例平均耗时(s)',
                       '是否劣化', '备注']
        wb = Workbook()
        ws = wb.active

        ws.column_dimensions['A'].width = 35.0
        ws.column_dimensions['B'].width = 12.0
        ws.column_dimensions['C'].width = 22.0
        ws.column_dimensions['D'].width = 22.0
        ws.column_dimensions['E'].width = 22.0
        ws.column_dimensions['F'].width = 22.0
        ws.column_dimensions['G'].width = 22.0
        ws.column_dimensions['H'].width = 20.0
        ws.column_dimensions['I'].width = 20.0
        ws.column_dimensions['J'].width = 20.0
        ws.column_dimensions['K'].width = 20.0
        ws.append(headers_row)
        ws.freeze_panes = 'A2'
        wb.save(daily_report_file)
    Constants.logger.info("initialize today report successfully")


def append_date_label(target_str, date_input):
    formatted_date = date_input.strftime('%Y%m%d')
    new_str = target_str + "_{}".format(formatted_date)

    return new_str


def get_given_date_report_name(date_input):
    report_name_head = "ffi_workload_daily_report"
    report_name_head = append_date_label(report_name_head, date_input)
    return report_name_head + ".xlsx"


def get_given_date_report_path(rerport_folder_path, date_input):
    report_file_name = get_given_date_report_name(date_input)
    report_file_path = os.path.join(rerport_folder_path, report_file_name)
    return report_file_path


if __name__ == "__main__":
    """
        command format: python3  get_ffi_workload_report.py  -i work_path -o report_folder_path
    """
    paras = get_args()
    work_path = paras.work_path
    log_path = os.path.join(work_path, "test.log")
    if os.path.exists(log_path):
        os.remove(log_path)

    Constants.logger = get_logger("workloadtest", log_path)
    Constants.logger.info("begin to get ffi workoad report.")
    Constants.logger.info("execute arguments: %s", paras)
    xml_path = os.path.join(work_path, "xmls")
    if not os.path.exists(xml_path) or not os.path.isdir(xml_path):
        Constants.logger.error("bad work parameter--work, please check it! work: %s", xml_path)
        sys.exit()
    today = datetime.date.today()
    yesterday = today - datetime.timedelta(days=1)
    today_daily_report_path = get_given_date_report_path(paras.rerport_folder_path, today)
    yesterday_daily_report_path = get_given_date_report_path(paras.rerport_folder_path, yesterday)

    if os.path.exists(today_daily_report_path):
        os.remove(today_daily_report_path)

    init_report(today_daily_report_path)
    today_test_data = get_today_test_result(xml_path)
    yestoday_test_data = get_yestoday_test_data(yesterday_daily_report_path)
    ret = generate_daily_report(today_daily_report_path, today_test_data, yestoday_test_data)
    if ret == Constants.RET_OK:
        summary_data_dict = get_summary_info(paras.work_path, today_daily_report_path)
        append_summary_info(today_daily_report_path, summary_data_dict)
        Constants.logger.info("run worklaod performance use cases finished. Please check details in report.")
    else:
        Constants.logger.info("run worklaod performance use cases failed. Please check details in log.")
