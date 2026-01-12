#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Copyright (c) 2024 Huawei Device Co., Ltd.
Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.

Description: Use ark to execute workload test suite
"""

import argparse
import datetime
import json
import logging
import os
import shutil
import stat
import subprocess
import sys
from typing import Union, List, Tuple
from collections import namedtuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


def get_logger(logger_name, log_file_path, level=logging.INFO):
    """Create logger for this script"""
    formatter = logging.Formatter(fmt='[%(asctime)s]  [%(levelname)s]   %(message)s',
                                  datefmt='%Y-%m-%d  %H:%M:%S')

    fh = logging.FileHandler(encoding='utf-8', mode='a', filename=log_file_path)
    fh.setFormatter(formatter)
    fh.setLevel(logging.DEBUG)
    # console output
    ch = logging.StreamHandler()
    ch.setFormatter(formatter)
    ch.setLevel(logging.DEBUG)
    log = logging.getLogger(logger_name)
    log.addHandler(fh)
    log.addHandler(ch)
    log.setLevel(level)

    return log


class Constants:
    logger = None
    CUR_PATH = os.path.abspath(os.path.dirname(__file__))
    TMP_PATH = os.path.join(os.getcwd(), "tmp")
    REPORT_NAME_HEAD_FIX = "js_perf_test_result"
    RET_OK = 0
    RET_FAILED = 1
    BINARY_PATH = ""
    OUTPUT_PATH = ""
    LOG_PATH = ""
    TODAY_EXCEL_PATH = ""
    YESTERDAY_EXCEL_PATH = ""
    DETERIORATION_BOUNDARY_VALUE = 0.05
    TODAY_EXCUTE_INFO = {}
    YESTERDAY_EXCUTE_TIME_DICT = {}
    V_8_EXCUTE_TIME_DICT = {}
    V_8_JITLESS_EXCUTE_TIME_DICT = {}
    JS_FILE_SUPER_LINK_DICT = {}
    HYPERLINK_HEAD = "https://gitee.com/dov1s/arkjs-perf-test/tree/builtins_test1110/js-perf-test"
    PASS = 'pass'
    FAIL = 'fail'
    SOLID = 'solid'
    NA_FIX = 'NA'
    # 1e-6 s
    COMPARISON_ACCURACY = 0.001
    ICU_DATA_PATH = ""
    FIX_STR = "8/d"
    V_8_ENGINED_PATH = '/usr/bin/v{}8'.format(FIX_STR)
    VER_PLATFORM = "full_x86_64"
    ES2ABC_PATH = ""
    ARK_JS_VM_PATH = ""
    ETS_RUNTIME = ""
    LD_LIBRARY_PATH = ""
    HDC_PATH: str = "hdc"
    DEVICE_WORKDIR: str = "/data/local/tmp/jsperftest"
    LIBS_LIST: List[str] = []
    STUB_AN: str = ""
    TASKSET_MASK: str = ""
    CaseTestDataType = namedtuple('test', ['exec_status', 'exec_time'])


def get_js_file_class_api_scenes(js_file_path):
    """Get all cases for one benchmark file"""
    scenes = []
    with open(js_file_path, 'r') as f:
        for line in f:
            if "scene_output" in line:
                str_array = line.split(':')
                mid_str = str_array[1].strip()
                elements = mid_str.split(' ')
                main_key = '/'.join([elements[0], elements[1] + '.js', elements[2]]).lower()
                scenes.append(main_key)
    return scenes


def degraded_str(yesterday_excute_time, exec_time):
    is_degraded_str = Constants.NA_FIX
    if len(str(yesterday_excute_time).strip()) != 0:
        if abs(float(yesterday_excute_time)) <= Constants.COMPARISON_ACCURACY:
            is_degraded_str = str(True) if abs(float(exec_time)) >= DETERIORATION_BOUNDARY_VALUE else str(False)
        else:
            is_degraded_tmp = float(exec_time) / float(yesterday_excute_time) >= (1 + DETERIORATION_BOUNDARY_VALUE)
            is_degraded_str = str(True) if is_degraded_tmp else str(False)

    return is_degraded_str


def v_8_excute_time_compute(main_key):
    v_8_excute_time_str = ''
    if len(Constants.V_8_EXCUTE_TIME_DICT) > 0 and main_key in Constants.V_8_EXCUTE_TIME_DICT.keys():
        v_8_excute_time_str = Constants.V_8_EXCUTE_TIME_DICT[main_key].strip()

    if len(v_8_excute_time_str) == 0:
        v_8_excute_time = ' '
    else:
        v_8_excute_time = v_8_excute_time_str

    return v_8_excute_time


def v_8_gitless_excute_time_compute(main_key):
    v_8_jitless_excute_time_str = ''
    if len(Constants.V_8_JITLESS_EXCUTE_TIME_DICT) > 0 and main_key in Constants.V_8_JITLESS_EXCUTE_TIME_DICT.keys():
        v_8_jitless_excute_time_str = Constants.V_8_JITLESS_EXCUTE_TIME_DICT[main_key].strip()

    if len(v_8_jitless_excute_time_str) == 0:
        v_8_jitless_excute_time = ' '
    else:
        v_8_jitless_excute_time = v_8_jitless_excute_time_str

    return v_8_jitless_excute_time


def ark_divide_v_8_compute(exec_time, v_8_excute_time):
    if len(exec_time) == 0 or len(v_8_excute_time.strip()) == 0:
        ark_divide_v_8 = Constants.NA_FIX
    elif abs(float(exec_time)) <= Constants.COMPARISON_ACCURACY:
        if abs(float(v_8_excute_time)) <= Constants.COMPARISON_ACCURACY:
            ark_divide_v_8 = '1'
        else:
            ark_divide_v_8 = '0'
    else:
        v_8_excute_time = v_8_excute_time.strip()
        if len(v_8_excute_time) == 0 or abs(float(v_8_excute_time)) <= Constants.COMPARISON_ACCURACY:
            ark_divide_v_8 = Constants.NA_FIX
        else:
            ark_divide_v_8 = str("{:.2f}".format(float(exec_time) / float(v_8_excute_time)))

    return ark_divide_v_8


def cast_to_float_or_str(value: str) -> Union[float, str]:
    """Return float value by str if it is possible, return input str otherwise"""
    try:
        result = float(value)
    except ValueError:
        result = value
    return result


def append_row_data(report_file, case_test_data):
    wb = load_workbook(report_file)
    ws = wb.worksheets[0]
    for main_key in case_test_data.keys():
        str_arr = main_key.split('/')
        class_name = str_arr[0]
        api_name = str_arr[1]
        scene = str_arr[2]
        js_case_name = '/'.join([class_name, api_name])
        excute_status = case_test_data[main_key].exec_status
        exec_time = case_test_data[main_key].exec_time.strip()
        yesterday_excute_time = ''
        if (len(Constants.YESTERDAY_EXCUTE_TIME_DICT) > 0 and
                Constants.YESTERDAY_EXCUTE_TIME_DICT.get(main_key) is not None):
            yesterday_excute_time = str(Constants.YESTERDAY_EXCUTE_TIME_DICT[main_key])
        is_degraded_str = degraded_str(yesterday_excute_time, exec_time)
        v_8_excute_time = v_8_excute_time_compute(main_key)
        v_8_jitless_excute_time = v_8_gitless_excute_time_compute(main_key)
        ark_divide_v_8 = ark_divide_v_8_compute(exec_time, v_8_excute_time)
        if len(exec_time) == 0 or len(v_8_jitless_excute_time.strip()) == 0:
            ark_divide_v_8_with_jitless = Constants.NA_FIX
        elif abs(float(exec_time)) <= Constants.COMPARISON_ACCURACY:
            if abs(float(v_8_jitless_excute_time)) <= Constants.COMPARISON_ACCURACY:
                ark_divide_v_8_with_jitless = '1'
            else:
                ark_divide_v_8_with_jitless = '0'
        else:
            v_8_jitless_excute_time = v_8_jitless_excute_time.strip()
            if (len(v_8_jitless_excute_time) == 0 or
                    abs(float(v_8_jitless_excute_time)) <= Constants.COMPARISON_ACCURACY):
                ark_divide_v_8_with_jitless = Constants.NA_FIX
            else:
                ark_divide_v_8_with_jitless = str("{:.2f}".format(float(exec_time) / float(v_8_jitless_excute_time)))
        jis_case_file_name_with_class = Constants.JS_FILE_SUPER_LINK_DICT['/'.join([class_name, api_name])]
        js_file_super_link = '/'.join([Constants.HYPERLINK_HEAD, jis_case_file_name_with_class])
        new_row = [js_case_name, scene, excute_status, cast_to_float_or_str(exec_time), yesterday_excute_time,
                   is_degraded_str, cast_to_float_or_str(v_8_excute_time),
                   cast_to_float_or_str(v_8_jitless_excute_time), cast_to_float_or_str(ark_divide_v_8),
                   cast_to_float_or_str(ark_divide_v_8_with_jitless), js_file_super_link, ' ']
        ws.append(new_row)
        check(is_degraded_str, ark_divide_v_8, ark_divide_v_8_with_jitless, ws)
    wb.save(report_file)
    return Constants.RET_OK


def check(is_degraded_str, ark_divide_v_8, ark_divide_v_8_with_jitless, ws):
    if is_degraded_str is str(True):
        ws.cell(row=ws.max_row, column=6).fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                                             fill_type=Constants.SOLID)
    if (ark_divide_v_8 != Constants.NA_FIX and
            (float(ark_divide_v_8) > 2 or abs(float(ark_divide_v_8) - 2) <= Constants.COMPARISON_ACCURACY)):
        ws.cell(row=ws.max_row, column=9).fill = PatternFill(start_color='FFFF00', end_color='FFFF00',
                                                             fill_type=Constants.SOLID)
    if (ark_divide_v_8_with_jitless != Constants.NA_FIX and
            (float(ark_divide_v_8_with_jitless) > 2 or
             abs(float(ark_divide_v_8_with_jitless) - 2) <= Constants.COMPARISON_ACCURACY)):
        ws.cell(row=ws.max_row, column=10).fill = PatternFill(start_color='FF00FF', end_color='FF00FF',
                                                              fill_type=Constants.SOLID)


def get_ark_js_cmd(abc_file: str) -> List[str]:
    """Get command for ark js vm"""
    cmd: List[str] = []
    if Constants.VER_PLATFORM.find("arm64") != -1:
        cmd = [Constants.HDC_PATH, "shell"]
        run_cmd = f"LD_LIBRARY_PATH={Constants.LD_LIBRARY_PATH}"
        if len(Constants.TASKSET_MASK) != 0:
            run_cmd += f" taskset -a {Constants.TASKSET_MASK}"
        run_cmd += " " + os.path.join(Constants.DEVICE_WORKDIR, "ark_js_vm")
        run_cmd += " --stub-file " + os.path.join(Constants.DEVICE_WORKDIR, "lib", "stub.an")
        run_cmd += " --icu-data-path " + os.path.join(Constants.DEVICE_WORKDIR, "data")
        run_cmd += " " + abc_file
        cmd.append(run_cmd)
    else:
        cmd = [Constants.ARK_JS_VM_PATH,
               "--log-level=info",
               "--enable-runtime-stat=true",
               "--stub-file", Constants.STUB_AN,
               "--icu-data-path", ICU_DATA_PATH,
               abc_file]
        if len(Constants.TASKSET_MASK) != 0:
            cmd = ["taskset", "-a", Constants.TASKSET_MASK] + cmd
    return cmd


def prepare_for_ark_run(class_name: str, api_name: str) -> Tuple[str, str]:
    """Prepare workspace for benchmark"""
    fangzhou_test_path = os.path.join(Constants.TMP_PATH, "fangzhou_test")  # for abc file
    if os.path.exists(fangzhou_test_path):
        shutil.rmtree(fangzhou_test_path)
    os.makedirs(fangzhou_test_path)
    class_folder_path = os.path.join(fangzhou_test_path, class_name)
    if not os.path.exists(class_folder_path):
        os.makedirs(class_folder_path)
    return (os.path.join(Constants.CUR_PATH, api_name + ".abc"),
            os.path.join(class_folder_path, api_name + ".log"))


def run_es2panda(abc_file: str, js_file: str) -> int:
    """Run es2panda for one benchmark file"""
    cmd = [Constants.ES2ABC_PATH, "--output", abc_file, js_file]
    logger.info("run cmd: %s", cmd)
    ret = subprocess.run(cmd, check=False)
    if ret.returncode != 0:
        logger.error("ret = %s, %s generate abc file failed. cmd: %s", str(ret), js_file, cmd)
    return ret.returncode


def run_js_case_via_ark(js_file_path, class_name, api_name, iterations, report_file):
    """Run js perf benchmark via ark js vm"""
    composite_scenes = get_js_file_class_api_scenes(js_file_path)
    case_test_data = {}
    execute_status = Constants.FAIL
    execute_time = ' '

    for _, composite_scene in enumerate(composite_scenes):
        case_test_data[composite_scene] = Constants.CaseTestDataType(execute_status, execute_time)

    js_file_name = class_name + '/' + api_name + '.js'
    cur_abc_file, api_log_path = prepare_for_ark_run(class_name, api_name)
    using_abc_file = cur_abc_file

    ret = run_es2panda(cur_abc_file, js_file_path)
    if ret != 0:
        append_row_data(report_file, case_test_data)
        return case_test_data
    if Constants.VER_PLATFORM.find("arm64") != -1:
        using_abc_file = os.path.join(Constants.DEVICE_WORKDIR, os.path.basename(cur_abc_file))
        if hdc_send(cur_abc_file, using_abc_file) != 0:
            append_row_data(report_file, case_test_data)
            return case_test_data
    # execute abc
    cmd = get_ark_js_cmd(using_abc_file)

    logger.info("run cmd: %s", cmd)
    flags = os.O_WRONLY | os.O_CREAT | os.O_EXCL
    modes = stat.S_IWUSR | stat.S_IRUSR

    data = {}
    for _ in range(iterations):
        if os.path.exists(api_log_path):
            os.remove(api_log_path)
        with os.fdopen(os.open(api_log_path, flags, modes), 'wb') as outfile:
            ret = subprocess.run(cmd, stdout=outfile)

        if ret.returncode != 0:
            logger.error("%s execute abc file failed. cmd: %s", js_file_name, cmd)
            append_row_data(report_file, case_test_data)
            return case_test_data
        if os.path.exists(api_log_path):
            data = update_data_by_log(data, api_log_path, js_file_name[:-3])

    case_test_data.clear()
    execute_status = Constants.PASS
    for k, time_value in data.items():
        case_test_data[k] = Constants.CaseTestDataType(execute_status, str(time_value / iterations))
    append_row_data(report_file, case_test_data)
    logger.info("%s execute abc file successfully. cmd: %s case_test_data: %s",
                js_file_name, cmd, case_test_data)
    os.remove(cur_abc_file)
    return case_test_data


def run_via_ark(jspath, report_file, iterations):
    if not os.path.exists(jspath):
        logger.error("js perf cases path is not exist. jspath: %s", jspath)
    logger.info("begin to run js perf test via ark. js perf cases path: %s", jspath)
    for root, _, files in os.walk(jspath):
        if "TestCaseError" in root:
            continue
        for file in files:
            if not file.endswith('.js'):
                continue

            file_path = os.path.join(root, file)
            results = file_path.split("/")
            class_name = results[-2]
            api_name = results[-1].split(".")[0]
            js_case_name = '/'.join([class_name, results[-1]])
            logger.info("begin to execute %s.", js_case_name)
            test_data = run_js_case_via_ark(file_path, class_name, api_name, iterations, report_file)
            for _, key in enumerate(test_data.keys()):
                Constants.TODAY_EXCUTE_INFO[key] = test_data.get(key)
            logger.info("finish executing %s. executing info: %s.", js_case_name, Constants.TODAY_EXCUTE_INFO)


def get_js_case_super_link_data(jspath):
    logger.info("get js case super link data")
    for root, _, files in os.walk(jspath):
        for file in files:
            if not file.endswith('.js'):
                continue

            file_path = os.path.join(root, file)
            results = file_path.split("/")
            class_name = results[-2]
            js_case_name = '/'.join([class_name, results[-1]])
            key = js_case_name.lower()
            Constants.JS_FILE_SUPER_LINK_DICT[key] = js_case_name


def export_sumary_info_for_notifying_email(json_path, total_cases_num, ark_divide_v_8_num, ark_divide_v_8_jitless_num):
    data = {}
    data['kind'] = 'V 8 js-perf-test'
    data['Total'] = total_cases_num
    data['Ark劣化v 8'] = ark_divide_v_8_num
    data['Ark劣化v 8 jitless'] = ark_divide_v_8_jitless_num
    flags = os.O_WRONLY | os.O_CREAT | os.O_EXCL
    modes = stat.S_IWUSR | stat.S_IRUSR
    if os.path.exists(json_path):
        os.remove(json_path)
    with os.fdopen(os.open(json_path, flags, modes), 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4, ensure_ascii=False)
        logger.info("export summary info to json file successfully.")


def get_umary_info_json_file_path(daily_report_file_path):
    dir_path = os.path.dirname(daily_report_file_path)
    json_file_name = 'jsperftest_notifying_info_in_email.json'
    json_file_path = os.path.join(dir_path, json_file_name)
    return json_file_path


def append_summary_info(report_file, total_cost_time):
    """
        summary info:
            pass count:
            fail count:
            totle count:
            degraded count:
            total excute time is(s) :
            degraded percentage upper limit:
            ark/v 8 degraded count:
            ark/v 8 jitless degraded count:
    """
    wb = load_workbook(report_file)
    ws = wb.worksheets[0]

    totle_num = 0
    degraded_upper_limit = DETERIORATION_BOUNDARY_VALUE
    pass_num = 0
    failed_num = 0
    degraded_num = 0
    ark_divide_v_8_degraded_count = 0
    ark_divide_v_8_jitless_degraded_count = 0

    last_bench_line = ws.max_row
    for row_num in range(2, last_bench_line + 1):
        excu_status = str(ws.cell(row=row_num, column=3).value)
        is_degraded = str(ws.cell(row=row_num, column=6).value)
        if is_degraded == str(True):
            degraded_num += 1

        if excu_status == Constants.PASS:
            pass_num += 1
            totle_num += 1
        elif excu_status == Constants.FAIL:
            failed_num += 1
            totle_num += 1

        obj = ws.cell(row=row_num, column=9).value
        if obj is None:
            obj = 0
        ark_divide_v_8 = obj
        if ark_divide_v_8 != Constants.NA_FIX and float(ark_divide_v_8) > 1:
            ark_divide_v_8_degraded_count += 1
        obj = ws.cell(row=row_num, column=10).value
        if obj is None:
            obj = 0
        ark_divide_v_8_jitless = obj
        if ark_divide_v_8_jitless != Constants.NA_FIX and float(ark_divide_v_8_jitless) > 1:
            ark_divide_v_8_jitless_degraded_count += 1

    avg_funcs = ['AVERAGE', 'GEOMEAN', 'MEDIAN']
    for avg_func in avg_funcs:
        new_row = [' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ',
                   f'=\"{avg_func}: \"&{avg_func}(I2:I{last_bench_line})',
                   f'=\"{avg_func}: \"&{avg_func}(J2:J{last_bench_line})', ' ', ' ']
        ws.append(new_row)
    new_row = ['劣化判定比率上限', degraded_upper_limit, ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
    ws.append(new_row)
    new_row = ['js 用例总数', totle_num, ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
    ws.append(new_row)
    new_row = ['Pass 数量', pass_num, ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
    ws.append(new_row)
    new_row = ['Fail 数量', failed_num, ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
    ws.append(new_row)
    new_row = ['ark今日劣化数量', degraded_num, ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
    ws.append(new_row)
    new_row = ['Total excute time(时:分:秒.微妙)', total_cost_time, ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
    ws.append(new_row)
    new_row = ['ark/v 8 劣化数量', ark_divide_v_8_degraded_count, ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ']
    ws.append(new_row)
    new_row = ['ark/v 8 jitless 劣化数量', ark_divide_v_8_jitless_degraded_count, ' ', ' ', ' ', ' ', ' ', ' ', ' ', ' ',
               ' ', ' ']
    ws.append(new_row)

    ws.column_dimensions.group('E', hidden=True)
    wb.save(report_file)

    json_file_path = get_umary_info_json_file_path(report_file)
    export_sumary_info_for_notifying_email(json_file_path, totle_num, ark_divide_v_8_degraded_count,
                                           ark_divide_v_8_jitless_degraded_count)
    return Constants.RET_OK


def process_args(args: argparse.Namespace) -> argparse.Namespace:
    """Process and check argument values"""
    if not os.path.exists(args.binarypath):
        logger.error("parameter --binarypath is not exist. Please check it! binary path: %s", args.binarypath)
        raise RuntimeError("error bad  parameters  --binarypath")
    if args.output_folder_path is None:
        args.output_folder_path = os.getcwd()
    if not os.path.isabs(args.output_folder_path):
        args.output_folder_path = os.path.abspath(args.output_folder_path)
    if args.ver_platform.find("arm64") == -1 and not os.path.exists(args.d_8_binary_path):
        logger.error("parameter --d_8_binary_path is not exist. Please check it! d 8  binary path: %s",
                     args.d_8_binary_path)
        raise RuntimeError("error bad  parameters  --d_8_binary_path: {}".format(args.d_8_binary_path))
    if args.iterations <= 0:
        logger.error("parameter --iterations <= 0. Please check it! iterations: %s",
                     args.iterations)
        raise RuntimeError(f"error bad  parameters --iterations: {args.iterations}")
    return args


def get_args():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--binarypath", "-bp", required=True,
        help="path of binary folder. refer to harmony root folder path",
    )
    parser.add_argument(
        "--jspath", "-p", required=True,
        help="path of js scripts, support folder and file",
    )
    parser.add_argument(
        "--deterioration_boundary_value", "-d", default=0.05,
        help="deterioration boundary value, default 0.05",
    )
    parser.add_argument(
        "--output_folder_path", "-o", default=None,
        help="output folder for executing js cases, default current folder",
    )
    parser.add_argument(
        "--d_8_binary_path", "-v", default=None,
        help="v 8 engine d 8 binary path",
    )
    parser.add_argument(
        "--ver_platform", "-e", default="full_x86_64",
        help="Code repository version and platform",
    )
    parser.add_argument(
        "--iterations", "-n", default=1, type=int,
        help="Number of benchmark launches"
    )
    parser.add_argument(
        "--hdc", default="hdc", type=str,
        help="path to hdc"
    )
    parser.add_argument(
        "--taskset", "-t", default="", type=str,
        help="Use taskset mask for affinity on specific CPUs"
    )
    parser.add_argument("--config", "-c", required=True, type=str, help="config json-file")
    return process_args(parser.parse_args())


def init_report(report_file):
    try:
        today_wb = load_workbook(report_file)
        today_ws = today_wb.worksheets[0]
    except FileNotFoundError:
        headers_row = ['用例名称', '场景', '执行状态', 'ark用例执行耗时(ms)', '昨日ark用例执行耗时(ms)', '是否劣化',
                       'v 8(ms)', 'v 8 --jitless(ms)', 'ark/v 8', 'ark/v 8 jitless', 'hyperlink', '备注']
        today_wb = Workbook()
        today_ws = today_wb.active

        today_ws.column_dimensions['A'].width = 35.0
        today_ws.column_dimensions['B'].width = 15.0
        today_ws.column_dimensions['C'].width = 15.0
        today_ws.column_dimensions['D'].width = 15.0
        today_ws.column_dimensions['E'].width = 25.0
        today_ws.column_dimensions['F'].width = 15.0
        today_ws.column_dimensions['G'].width = 15.0
        today_ws.column_dimensions['H'].width = 15.0
        today_ws.column_dimensions['I'].width = 15.0
        today_ws.column_dimensions['J'].width = 15.0
        today_ws.column_dimensions['K'].width = 50.0
        today_ws.column_dimensions['L'].width = 15.0
        today_ws.append(headers_row)
        today_ws.freeze_panes = 'A2'
        today_wb.save(report_file)


def append_date_label(target_str, date_input):
    formatted_date = date_input.strftime('%Y%m%d')
    new_str = target_str + "_{}".format(formatted_date)

    return new_str


def get_v_8_benchmark_daily_report_path():
    '''
        get v 8 based data. v 8 based data obtained on 1,11,21 day for dayevery month.that is to say, in 1,11,21,
        v 8 executes js cases.
    '''
    now = datetime.datetime.now(tz=datetime.timezone.utc)
    today_str = now.strftime("%Y.%m.%d")
    str_list = today_str.split('.')
    year_str = str_list[0]
    month_str = str_list[1]
    day = int(str_list[2])
    based_day = 0
    if day > 21:
        based_day = 21
    elif day > 11:
        based_day = 11
    else:
        based_day = 1

    based_date = year_str + month_str + str(based_day)
    base_date_file = based_date + '.xlsx'
    based_report_name = '_'.join([Constants.REPORT_NAME_HEAD_FIX, base_date_file])
    report_file_path = os.path.join(OUTPUT_PATH, based_report_name)
    return report_file_path


def get_given_date_report_name(date_input):
    report_name_head = append_date_label(Constants.REPORT_NAME_HEAD_FIX, date_input)
    return report_name_head + ".xlsx"


def get_given_date_report_path(date_input):
    report_file_name = get_given_date_report_name(date_input)
    report_file_path = os.path.join(OUTPUT_PATH, report_file_name)
    return report_file_path


def get_yesterday_excute_times(yesterday_report):
    if not os.path.exists(yesterday_report) or not os.path.isfile(yesterday_report):
        return

    wb = load_workbook(yesterday_report)
    ws = wb.worksheets[0]
    for row_num in range(2, ws.max_row + 1):
        js_case = ws.cell(row=row_num, column=1).value
        scene = ws.cell(row=row_num, column=2).value
        exec_status = ws.cell(row=row_num, column=3).value
        if exec_status == Constants.PASS or exec_status == Constants.FAIL:
            main_key = '/'.join([js_case, scene]).lower()
            excute_time = ws.cell(row=row_num, column=4).value
            Constants.YESTERDAY_EXCUTE_TIME_DICT[main_key] = excute_time


def update_data_by_log(data: dict, log_path: str, js_name: str) -> dict:
    """Update execution time data by log file"""
    with open(log_path, 'r') as f:
        for line in f:
            if "scene_output" not in line:
                continue
            str_array = line.split(':')
            mid_str = str_array[1].strip()
            scene = mid_str.split()[2]
            exec_time = str_array[2]
            key_str = '/'.join([js_name + '.js', scene]).lower()
            if key_str not in data:
                data[key_str] = float(exec_time)
            else:
                data[key_str] += float(exec_time)
    return data


def get_v_8_cmd(parameter: str, js_file_path: str) -> List[str]:
    """Get command for v 8"""
    cmd: List[str] = []
    if Constants.VER_PLATFORM.find("arm64") != -1:
        cmd = [Constants.HDC_PATH, "shell"]
    if len(Constants.TASKSET_MASK) != 0:
        cmd += ["taskset", "-a", Constants.TASKSET_MASK]
    if len(parameter) == 0:
        cmd += [Constants.V_8_ENGINED_PATH, js_file_path]
    else:
        cmd += [Constants.V_8_ENGINED_PATH, parameter, js_file_path]
    return cmd


def run_v_8_single_js_case(js_file_path, cmd_para, js_case_name, iterations: int):
    """Run single js case for v 8 based with parameters"""
    v_8_exec_time_dict = {}
    scenes = get_js_file_class_api_scenes(js_file_path)

    v_8_log_path = os.path.join(Constants.CUR_PATH, "v_8.log")

    flags = os.O_WRONLY | os.O_CREAT | os.O_EXCL
    modes = stat.S_IWUSR | stat.S_IRUSR
    used_js_file = js_file_path

    if Constants.VER_PLATFORM.find("arm64") != -1:
        js_file_path_device = os.path.join(Constants.DEVICE_WORKDIR,
                                           os.path.basename(js_case_name + '.js'))
        used_js_file = js_file_path_device
        if hdc_send(js_file_path, js_file_path_device) != 0:
            for elem in enumerate(scenes):
                v_8_exec_time_dict[elem] = 0
            logger.error("Couldn't send file %s to device", js_file_path)
            return v_8_exec_time_dict

    cmd = get_v_8_cmd(cmd_para, used_js_file)
    logger.info("run cmd:%s", cmd)
    data = {}
    for _ in range(iterations):
        if os.path.exists(v_8_log_path):
            os.remove(v_8_log_path)
        with os.fdopen(os.open(v_8_log_path, flags, modes), 'wb') as outfile:
            ret = subprocess.run(cmd, stdout=outfile, check=False)
        if ret.returncode != 0:
            for elem in enumerate(scenes):
                v_8_exec_time_dict[elem] = 0
            logger.error("execute cmd failed. cmd: %s", cmd)
            return v_8_exec_time_dict
        data = update_data_by_log(data, v_8_log_path, js_case_name)
        os.remove(v_8_log_path)

    for k, time_value in data.items():
        v_8_exec_time_dict[k] = str(time_value / iterations)
    logger.info("v 8 excute %s successfully. cmd: %s", js_file_path, cmd)
    return v_8_exec_time_dict


def get_given_column_data(report_file, column_index):
    column_data = {}
    if os.path.exists(report_file) and report_file.endswith('.xlsx'):
        wb = load_workbook(report_file)
        ws = wb.worksheets[0]

        for row_num in range(2, ws.max_row + 1):
            js_case_name = str(ws.cell(row=row_num, column=1).value)
            scene = str(ws.cell(row=row_num, column=2).value)
            exec_status = str(ws.cell(row=row_num, column=3).value)
            time = str(ws.cell(row=row_num, column=column_index).value)
            if exec_status == Constants.PASS or exec_status == Constants.FAIL:
                main_key = '/'.join([js_case_name, scene])
                column_data[main_key] = time

    return column_data


def get_v_8_excute_times(jspath, v_8_based_report_file, iterations):
    if os.path.exists(v_8_based_report_file) and os.path.isfile(v_8_based_report_file):
        # Generate v 8 benchmark data on the 1st, 11th, and 21st of each month.The testing at other times refers to
        # these V 8 benchmark data
        v_8_exec_time_dict = get_given_column_data(v_8_based_report_file, 7)
        for key in v_8_exec_time_dict.keys():
            Constants.V_8_EXCUTE_TIME_DICT[key] = v_8_exec_time_dict[key]
        return Constants.RET_OK

    file_list = []
    for root, _, files in os.walk(jspath):
        for file in files:
            if not file.endswith('.js'):
                continue
            file_path = os.path.join(root, file)
            file_list.append(file_path)
    for _, file_path in enumerate(file_list):
        results = file_path.split("/")
        class_name = results[-2]
        api_name = results[-1].split(".")[0]
        js_case_name = '/'.join([class_name, api_name])

        v_8_exec_time_dict = run_v_8_single_js_case(file_path, '', js_case_name, iterations)
        for key in v_8_exec_time_dict.keys():
            Constants.V_8_EXCUTE_TIME_DICT[key] = v_8_exec_time_dict[key]

    return Constants.RET_OK


def get_v_8_jitless_excute_times(jspath, v_8_based_report_file_path, iterations):
    if os.path.exists(v_8_based_report_file_path) and os.path.isfile(v_8_based_report_file_path):
        # Generate v 8 benchmark data on the 1st, 11th, and 21st of each month.The testing at other times refers to
        # these V 8 benchmark data
        v_8_exec_time_dict = get_given_column_data(v_8_based_report_file_path, 8)
        for key in v_8_exec_time_dict.keys():
            Constants.V_8_JITLESS_EXCUTE_TIME_DICT[key] = v_8_exec_time_dict[key]
        return Constants.RET_OK

    file_list = []
    for root, _, files in os.walk(jspath):
        for file in files:
            if not file.endswith('.js'):
                continue
            file_path = os.path.join(root, file)
            file_list.append(file_path)

    for _, file_path in enumerate(file_list):
        results = file_path.split("/")
        class_name = results[-2]
        api_name = results[-1].split(".")[0]
        js_case_name = '/'.join([class_name, api_name])

        v_8_exec_time_dict = run_v_8_single_js_case(file_path, '--jitless', js_case_name, iterations)
        for key in v_8_exec_time_dict.keys():
            Constants.V_8_JITLESS_EXCUTE_TIME_DICT[key] = v_8_exec_time_dict[key]

    return Constants.RET_OK


def hdc_send(source: str, destination: str) -> int:
    """Run hdc send command"""
    hdc_cmd: List[str] = [Constants.HDC_PATH, "file", "send"]
    hdc_cmd += [source, destination]
    logger.info("run cmd: %s", hdc_cmd)
    return subprocess.run(hdc_cmd, check=False).returncode


def hdc_run(cmd: List[str]) -> int:
    """Run command on device via hdc shell"""
    hdc_cmd = [Constants.HDC_PATH, "shell"] + cmd
    return subprocess.run(hdc_cmd).returncode


def prepare_device():
    """Preapare device workdir for js perf testing"""
    if hdc_send(Constants.ARK_JS_VM_PATH, Constants.DEVICE_WORKDIR) != 0:
        logger.error("Couldn't send ark_js_vm to device")
        sys.exit(1)
    hdc_run(["chmod", "u+x", os.path.join(Constants.DEVICE_WORKDIR, "ark_js_vm")])
    arkjsvm_lib = os.path.join(Constants.ETS_RUNTIME, "libark_jsruntime.so")
    if hdc_send(arkjsvm_lib, os.path.join(Constants.DEVICE_WORKDIR, "lib")) != 0:
        logger.error("Couldn't send libark_jsruntime.so to device")
        sys.exit(1)
    if hdc_send(ICU_DATA_PATH, Constants.DEVICE_WORKDIR) != 0:
        logger.error("Couldn't send icu data to device")
        sys.exit(1)
    thirdparty_path = os.path.join(Constants.DEVICE_WORKDIR, "thirdparty")
    for lib in Constants.LIBS_LIST:
        if not os.path.isdir(lib):
            logger.error("Couldn't find lib from config %s", lib)
            sys.exit(1)
        if hdc_send(lib, thirdparty_path) != 0:
            logger.error("Couldn't send %s lib to device", lib)
            sys.exit(1)
    if hdc_send(Constants.STUB_AN, os.path.join(Constants.DEVICE_WORKDIR, "lib")) != 0:
        logger.error("Couldn't send %s file to device", Constants.STUB_AN)
        sys.exit(1)


def get_config(parameters: argparse.Namespace):
    """Get config from arguments and json file"""
    Constants.V_8_ENGINED_PATH = parameters.d_8_binary_path
    Constants.VER_PLATFORM = parameters.ver_platform
    Constants.HDC_PATH = parameters.hdc
    Constants.TASKSET_MASK = parameters.taskset
    with open(parameters.config, 'r', encoding='UTF-8') as f:
        json_data = json.load(f)

    Constants.ES2ABC_PATH = os.path.join(BINARY_PATH, json_data[Constants.VER_PLATFORM]["ES2ABC"])
    Constants.ETS_RUNTIME = os.path.join(BINARY_PATH,
                                    json_data[Constants.VER_PLATFORM]["ETS_RUNTIME_PATH"])
    Constants.ARK_JS_VM_PATH = os.path.join(Constants.ETS_RUNTIME, "ark_js_vm")
    Constants.STUB_AN = os.path.join(BINARY_PATH, json_data[Constants.VER_PLATFORM]["STUB_AN"])
    libs = json_data[Constants.VER_PLATFORM]["LIBS_LIST"]
    for lib in libs:
        Constants.LIBS_LIST.append(os.path.normpath(os.path.join(BINARY_PATH, lib)))
    if Constants.VER_PLATFORM.find("x86_64") != -1:
        old_ld_library_path = os.environ.get('LD_LIBRARY_PATH', '')
        Constants.LD_LIBRARY_PATH = Constants.ETS_RUNTIME + ":"
        if len(Constants.LIBS_LIST) != 0:
            Constants.LD_LIBRARY_PATH += ":".join(Constants.LIBS_LIST)
        if len(old_ld_library_path) != 0:
            Constants.LD_LIBRARY_PATH += f":{old_ld_library_path}"
        os.environ['LD_LIBRARY_PATH'] = Constants.LD_LIBRARY_PATH
    elif Constants.VER_PLATFORM.find("arm64") != -1:
        Constants.LD_LIBRARY_PATH = os.path.join(Constants.DEVICE_WORKDIR, "lib")
        for lib in Constants.LIBS_LIST:
            lib = os.path.normpath(lib)
            Constants.LD_LIBRARY_PATH += ":" +\
                os.path.join(Constants.DEVICE_WORKDIR, "thirdparty", os.path.basename(lib))


if __name__ == "__main__":
    LOG_PATH = os.path.join(Constants.TMP_PATH, "test.log")
    if os.path.exists(LOG_PATH):
        os.remove(LOG_PATH)
    logger = get_logger("jstest", LOG_PATH)

    paras = get_args()
    logger.info("execute arguments: %s", paras)

    DETERIORATION_BOUNDARY_VALUE = paras.deterioration_boundary_value
    BINARY_PATH = paras.binarypath
    ICU_DATA_PATH = os.path.join(BINARY_PATH, "third_party/icu/ohos_icu4j/data")
    OUTPUT_PATH = Constants.CUR_PATH
    get_config(paras)
    if not os.path.exists(Constants.ARK_JS_VM_PATH):
        logger.error("%s does not exist", Constants.ARK_JS_VM_PATH)
        sys.exit(1)
    if Constants.VER_PLATFORM.find("arm64") != -1:
        prepare_device()

    if paras.output_folder_path is not None:
        OUTPUT_PATH = paras.output_folder_path

    if not os.path.exists(OUTPUT_PATH):
        os.makedirs(OUTPUT_PATH)

    today = datetime.date.today()
    yesterday = today - datetime.timedelta(days=1)
    TODAY_EXCEL_PATH = get_given_date_report_path(today)
    YESTERDAY_EXCEL_PATH = get_given_date_report_path(yesterday)

    if os.path.exists(TODAY_EXCEL_PATH):
        os.remove(TODAY_EXCEL_PATH)

    get_js_case_super_link_data(paras.jspath)
    start_time = datetime.datetime.now(tz=datetime.timezone.utc)
    init_report(TODAY_EXCEL_PATH)
    get_yesterday_excute_times(YESTERDAY_EXCEL_PATH)
    v_8_based_report_path = get_v_8_benchmark_daily_report_path()
    get_v_8_excute_times(paras.jspath, v_8_based_report_path, paras.iterations)
    get_v_8_jitless_excute_times(paras.jspath, v_8_based_report_path, paras.iterations)

    run_via_ark(paras.jspath, TODAY_EXCEL_PATH, paras.iterations)
    end_time = datetime.datetime.now(tz=datetime.timezone.utc)

    totol_time = u"%s" % (end_time - start_time)
    append_summary_info(TODAY_EXCEL_PATH, totol_time)

    logger.info("run js perf test finished. Please check details in report.")
    shutil.rmtree(Constants.TMP_PATH)
